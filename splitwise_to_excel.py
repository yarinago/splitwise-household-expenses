#!/usr/bin/env python3
"""
Splitwise → Excel (All‑History, Incremental Upsert + Per‑Person Owed Columns)

What this script does
---------------------
• Uses OAuth2 token from environment variables to read ALL expenses for a configured group.
• Incrementally upserts by expense_id (re‑runs only add/update deltas).
• For each expense, pulls per‑user splits from the expense's **users[].owed_share** (not repayments),
  and writes how much each tracked member owes **on the Raw_Expenses sheet itself**.
• Builds per‑month category totals and per‑person monthly owed totals.
• Writes an Excel file with:
    - Raw_Expenses  → includes columns for how much each member **owes** per expense (from `owed_share`), with columns named using the members' display names + "_owes".
    - Raw_Shares    → optional, long form: one row per (expense, person) with owed_share
                      (disable with --no-raw-shares if you don't want it)
    - Monthly_By_Category
    - PerPerson_Month → how much each tracked member owes per month (from `owed_share`)
    - Charts        → interactive dashboard with summary tiles and charts

Configuration (Environment Variables)
--------------------------------------
REQUIRED (GitHub Secrets):
  SPLITWISE_CLIENT_ID            OAuth2 client ID
  SPLITWISE_CLIENT_SECRET        OAuth2 client secret
  SPLITWISE_ACCESS_TOKEN_JSON    OAuth2 token JSON (format: {"access_token":"...","token_type":"bearer","refresh_token":"..."})

REQUIRED (GitHub Variables):
  SPLITWISE_GROUP_ID             Splitwise group ID
  SPLITWISE_MEMBERS              JSON map of user IDs to names (e.g., {"user_id": "name", ...})

OPTIONAL (GitHub Variables):
  SPLITWISE_FIRST_MONTH          Start month for history (default: 2008-01)
  SPLITWISE_EXCLUDE_MONTHS       Months to exclude (CSV, e.g., 2025-10,2025-11)
  SPLITWISE_EXCLUDE_DESCRIPTIONS Description patterns to exclude (CSV, case-insensitive)

See README.md for detailed setup instructions.
Hebrew is fully supported in descriptions, categories, and names.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import sys
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from dotenv import load_dotenv

# ---------- Configuration (from environment/CLI) ----------
_group_id_str = os.getenv("SPLITWISE_GROUP_ID", "").strip()
DEFAULT_GROUP_ID = int(_group_id_str) if _group_id_str else 0
FIRST_MONTH = os.getenv("SPLITWISE_FIRST_MONTH", "2008-01")
OUTPUT_FILE_TEMPLATE = os.getenv("SPLITWISE_OUTPUT_FILE", f"splitwise_group_{DEFAULT_GROUP_ID}_all_history.xlsx")
GLOBAL_MEMBER_OWES_SUM: Dict[str, float] = {}

# Members mapping (id → display name), loaded from env or CLI
def _load_members() -> Dict[int, str]:
    members_json = os.getenv("SPLITWISE_MEMBERS")
    if members_json:
        try:
            return {int(k): v for k, v in json.loads(members_json).items()}
        except Exception as e:
            print(f"[!] Failed to parse SPLITWISE_MEMBERS JSON: {e}", file=sys.stderr)
            raise SystemExit("[!] Provide valid SPLITWISE_MEMBERS JSON")
    else:
        raise SystemExit("[!] SPLITWISE_MEMBERS not set. Provide JSON dict like '{\"12345\": \"Alice\", \"67890\": \"Bob\"}'")

MEMBER_ID_TO_NAME: Dict[int, str] = _load_members()

# ---------- Splitwise SDK ----------
try:
    from splitwise import Splitwise
    from splitwise.expense import Expense
    from splitwise.group import Group
except Exception:
    print("[!] Missing or outdated 'splitwise' package. Install/upgrade with: pip install -U splitwise", file=sys.stderr)
    raise


# ---------- Helpers ----------
def ym_today() -> str:
    return dt.date.today().strftime("%Y-%m")


def month_iter(start_ym: str, end_ym: str) -> List[str]:
    start = dt.datetime.strptime(start_ym, "%Y-%m")
    end = dt.datetime.strptime(end_ym, "%Y-%m")
    if end < start:
        raise ValueError("end month is before start month")
    months = []
    cur = start
    while cur <= end:
        months.append(cur.strftime("%Y-%m"))
        year = cur.year + (cur.month // 12)
        month = (cur.month % 12) + 1
        cur = cur.replace(year=year, month=month, day=1)
    return months


def month_bounds(ym: str) -> Tuple[str, str]:
    start = dt.datetime.strptime(ym, "%Y-%m").date().replace(day=1)
    next_month_year = start.year + (start.month // 12)
    next_month = (start.month % 12) + 1
    end_excl = dt.date(next_month_year, next_month, 1)
    return start.isoformat(), end_excl.isoformat()


def make_client(client_id: str, client_secret: str, token_json: Dict) -> "Splitwise":
    sw = Splitwise(client_id, client_secret)
    if "access_token" in token_json:  # OAuth2
        token_json.setdefault("token_type", "bearer")
        if hasattr(sw, "setOAuth2Token"):
            sw.setOAuth2Token(token_json)
        elif hasattr(sw, "setOAuth2AccessToken"):
            sw.setOAuth2AccessToken(token_json)
        else:
            sw.setAccessToken(token_json)
        return sw
    if "oauth_token" in token_json and "oauth_token_secret" in token_json:  # OAuth1 legacy
        sw.setAccessToken(token_json)
        return sw
    raise SystemExit("[!] Unrecognized token JSON for Splitwise.")


def fetch_group(sw: Splitwise, group_id: int) -> Group:
    g = sw.getGroup(group_id)
    if not g:
        raise ValueError(f"Group id {group_id} not found")
    return g


def fetch_expenses_all_history(sw: Splitwise, group_id: int, start_ym: str, end_ym: str) -> List[Expense]:
    expenses: List[Expense] = []
    for ym in month_iter(start_ym, end_ym):
        start_iso, end_iso = month_bounds(ym)
        offset = 0
        while True:
            chunk = sw.getExpenses(group_id=group_id, dated_after=start_iso, dated_before=end_iso, offset=offset)
            if not chunk:
                break
            expenses.extend(chunk)
            offset += len(chunk)
            if len(chunk) < 20:
                break
    return expenses


def _safe_attr(obj, getter_name: str, default=None):
    try:
        getter = getattr(obj, getter_name)
        return getter() if callable(getter) else getter
    except Exception:
        return default

def _coerce_float(val, default=0.0) -> float:
    if val is None:
        return float(default)
    try:
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace(',', '')
        if s == '':
            return float(default)
        return float(s)
    except Exception:
        return float(default)


def normalize_expenses(expenses: Iterable[Expense]) -> pd.DataFrame:
    """Turn Splitwise Expense objects into a clean dataframe (one row per expense)."""
    rows = []
    for e in expenses:
        if _safe_attr(e, "getDeletedAt"):
            continue
        date = _safe_attr(e, "getDate")
        updated = _safe_attr(e, "getUpdatedAt")
        try:
            cost = float(_safe_attr(e, "getCost", 0.0) or 0.0)
        except Exception:
            cost = 0.0
        currency = _safe_attr(e, "getCurrencyCode")
        category_name = _safe_attr(_safe_attr(e, "getCategory"), "getName") or "Uncategorized"
        desc = _safe_attr(e, "getDescription") or ""
        eid = _safe_attr(e, "getId")
        month = date[:7] if date else None
        rows.append({
            "date": date,
            "updated_at": updated,
            "month": month,
            "description": desc,
            "category": category_name,
            "amount": cost,
            "currency": currency,
            "expense_id": eid,
        })
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    # filter out payments/transfers if you prefer
    mask_payment = df["category"].str.contains("payment|transfer", case=False, na=False)
    df = df[~mask_payment].copy()
    if "expense_id" in df.columns:
        df["expense_id"] = pd.to_numeric(df["expense_id"], errors="coerce").astype("Int64")
    for c in ("date","updated_at","month","description","category","currency"):
        if c in df.columns:
            df[c] = df[c].astype(str)
    if "amount" in df.columns:
        df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0.0)
    return df


def normalize_person_shares(expenses: Iterable[Expense]) -> pd.DataFrame:
    """One row per (expense, user) using **users[].owed_share** as the amount that user pays.
    Robust across SDK versions and raw dict payloads.
    """
    rows = []
    for e in expenses:
        if _safe_attr(e, "getDeletedAt"):
            continue
        date = _safe_attr(e, "getDate")
        month = date[:7] if date else None
        eid = _safe_attr(e, "getId")

        # get users list from SDK object or dict
        users = None
        if hasattr(e, "getUsers") and callable(getattr(e, "getUsers")):
            users = e.getUsers()
        if users is None and hasattr(e, "users"):
            users = getattr(e, "users")
        if users is None:
            # last resort: object __dict__
            users = getattr(e, "__dict__", {}).get("users")
        users = users or []

        for u in users:
            uid = None
            owed = None
            display_name = None

            # If dict-like
            if isinstance(u, dict):
                uid = u.get("user_id")
                if uid is None and isinstance(u.get("user"), dict):
                    uid = u["user"].get("id")
                owed = u.get("owed_share") or u.get("owedShare") or u.get("owed")
                display_name = (u.get("user") or {}).get("first_name")
            else:
                # SDK object path(s)
                # A) ExpenseUser inherits from User → id is available via getId()/id
                getter = getattr(u, "getId", None)
                if callable(getter):
                    uid = getter()
                if uid is None and hasattr(u, "id") and not callable(getattr(u, "id")):
                    uid = getattr(u, "id")
                # B) Some builds expose getUserId()
                if uid is None:
                    getter = getattr(u, "getUserId", None)
                    if callable(getter):
                        uid = getter()
                # C) Nested user object
                if uid is None:
                    user_obj = getattr(u, "getUser", None)
                    if callable(user_obj):
                        user_obj = user_obj()
                    if user_obj is not None:
                        getter = getattr(user_obj, "getId", None)
                        if callable(getter):
                            uid = getter()
                        elif hasattr(user_obj, "id") and not callable(getattr(user_obj, "id")):
                            uid = getattr(user_obj, "id")
                        # name from nested user
                        for attr in ("getFirstName", "first_name"):
                            val = getattr(user_obj, attr, None)
                            if callable(val):
                                val = val()
                            if val is not None:
                                display_name = val
                                break
                # owed share (support different attribute styles)
                for attr in ("getOwedShare", "owed_share", "owedShare", "getOwed"):
                    val = getattr(u, attr, None)
                    if callable(val):
                        val = val()
                    if val is not None:
                        owed = val
                        break

            # Fallback: if still no name, try direct first-name on u
            if not display_name and not isinstance(u, dict):
                for attr in ("getFirstName", "first_name"):
                    val = getattr(u, attr, None)
                    if callable(val):
                        val = val()
                    if val is not None:
                        display_name = val
                        break

            uid = int(uid) if uid is not None else None
            owed_val = _coerce_float(owed, 0.0)
            # map to your fixed display names if possible
            name = MEMBER_ID_TO_NAME.get(uid, display_name if display_name else (str(uid) if uid is not None else "Unknown"))

            rows.append({
                "expense_id": eid,
                "user_id": uid,
                "person": str(name) if name is not None else "Unknown",
                "month": month,
                "owed_share": owed_val,
            })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df = df[df["month"].notna()].copy()
    if "expense_id" in df.columns:
        df["expense_id"] = pd.to_numeric(df["expense_id"], errors="coerce").astype("Int64")
    if "user_id" in df.columns:
        df["user_id"] = pd.to_numeric(df["user_id"], errors="coerce").astype("Int64")
    df["owed_share"] = pd.to_numeric(df["owed_share"], errors="coerce").fillna(0.0)
    return df
    df = df[df["month"].notna()].copy()
    if "expense_id" in df.columns:
        df["expense_id"] = pd.to_numeric(df["expense_id"], errors="coerce").astype("Int64")
    if "user_id" in df.columns:
        df["user_id"] = pd.to_numeric(df["user_id"], errors="coerce").astype("Int64")
    df["owed_share"] = pd.to_numeric(df["owed_share"], errors="coerce").fillna(0.0)
    return df
    df = df[df["month"].notna()].copy()
    if "expense_id" in df.columns:
        df["expense_id"] = pd.to_numeric(df["expense_id"], errors="coerce").astype("Int64")
    if "user_id" in df.columns:
        df["user_id"] = pd.to_numeric(df["user_id"], errors="coerce").astype("Int64")
    df["owed_share"] = pd.to_numeric(df["owed_share"], errors="coerce").fillna(0.0)
    return df


def widen_raw_with_member_owed(raw_df: pd.DataFrame, shares_df: pd.DataFrame, id_to_name: Dict[int, str]) -> pd.DataFrame:
    """Return Raw_Expenses with extra columns: <name>_owes for each configured member.
    We intentionally take users[].owed_share as the amount each member **owes** per expense.
    Handles stale columns and ensures numeric, rounded values.
    """
    out = raw_df.copy()

    # Drop any stale member columns (_owed / _paid / _owes and their _x/_y variants) before merging
    cols_to_drop: List[str] = []
    for display_name in id_to_name.values():
        for base in (f"{display_name}_owes", f"{display_name}_owed", f"{display_name}_paid"):
            for cand in (base, f"{base}_x", f"{base}_y"):
                if cand in out.columns:
                    cols_to_drop.append(cand)
    if cols_to_drop:
        out.drop(columns=list(set(cols_to_drop)), inplace=True, errors="ignore")

    if shares_df.empty or out.empty:
        # Ensure final owes columns exist even if no data
        for display_name in id_to_name.values():
            base = f"{display_name}_owes"
            if base not in out.columns:
                out[base] = 0.0
        return out

    # Build pivot from users[].owed_share → "owes"
    wanted_ids = list(id_to_name.keys())
    sub = shares_df[shares_df["user_id"].isin(wanted_ids)].copy()
    piv = sub.pivot_table(index="expense_id", columns="user_id", values="owed_share", aggfunc="sum", fill_value=0.0)
    piv = piv.reindex(columns=wanted_ids, fill_value=0.0)
    piv.columns = [f"{id_to_name.get(int(uid), str(uid))}_owes" for uid in piv.columns]

    # Merge
    out = out.merge(piv.reset_index(), on="expense_id", how="left")

    # Ensure final _owes columns exist and are numeric
    for display_name in id_to_name.values():
        base = f"{display_name}_owes"
        if base not in out.columns:
            out[base] = 0.0
        out[base] = pd.to_numeric(out[base], errors="coerce").fillna(0.0).round(2)

    return out

    # Build pivot from users[].owed_share → we treat as "paid"
    wanted_ids = list(id_to_name.keys())
    sub = shares_df[shares_df["user_id"].isin(wanted_ids)].copy()
    piv = sub.pivot_table(index="expense_id", columns="user_id", values="owed_share", aggfunc="sum", fill_value=0.0)
    piv = piv.reindex(columns=wanted_ids, fill_value=0.0)
    piv.columns = [f"{id_to_name.get(int(uid), str(uid))}_paid" for uid in piv.columns]

    # Merge
    out = out.merge(piv.reset_index(), on="expense_id", how="left")

    # Ensure final _paid columns exist and are numeric
    for display_name in id_to_name.values():
        base = f"{display_name}_paid"
        if base not in out.columns:
            out[base] = 0.0
        out[base] = pd.to_numeric(out[base], errors="coerce").fillna(0.0).round(2)

    return out

    # Build a compact pivot of owed shares per expense for the tracked ids
    wanted_ids = list(id_to_name.keys())
    sub = shares_df[shares_df["user_id"].isin(wanted_ids)].copy()
    piv = sub.pivot_table(index="expense_id", columns="user_id", values="owed_share", aggfunc="sum", fill_value=0.0)
    piv = piv.reindex(columns=wanted_ids, fill_value=0.0)  # fixed order
    piv.columns = [f"{id_to_name.get(int(uid), str(uid))}_owed" for uid in piv.columns]

    # Merge in owed columns (safe now that we dropped stale ones)
    out = out.merge(piv.reset_index(), on="expense_id", how="left")

    # Ensure owed columns exist and are numeric
    for display_name in id_to_name.values():
        base = f"{display_name}_owed"
        if base not in out.columns:
            out[base] = 0.0
        out[base] = pd.to_numeric(out[base], errors="coerce").fillna(0.0).round(2)

    return out

    # Build a compact pivot of owed shares per expense for the tracked ids
    wanted_ids = list(id_to_name.keys())
    sub = shares_df[shares_df["user_id"].isin(wanted_ids)].copy()
    piv = sub.pivot_table(index="expense_id", columns="user_id", values="owed_share", aggfunc="sum", fill_value=0.0)
    piv = piv.reindex(columns=wanted_ids, fill_value=0.0)  # fixed order
    piv.columns = [f"{id_to_name.get(int(uid), str(uid))}_owed" for uid in piv.columns]

    # Merge in owed columns; if columns already exist, pandas will create _x/_y
    out = out.merge(piv.reset_index(), on="expense_id", how="left")

    # Coalesce any *_owed_x/_owed_y into a single *_owed, then clean types
    for display_name in id_to_name.values():
        base = f"{display_name}_owed"
        cx, cy = f"{base}_x", f"{base}_y"
        if cy in out.columns:
            out[base] = out[cy]
        elif cx in out.columns and base not in out.columns:
            out[base] = out[cx]
        if base not in out.columns:
            out[base] = 0.0
        out[base] = pd.to_numeric(out[base], errors="coerce").fillna(0.0).round(2)
        # Drop the temporary suffix columns if present
        if cx in out.columns:
            out.drop(columns=[cx], inplace=True)
        if cy in out.columns:
            out.drop(columns=[cy], inplace=True)

    return out
    wanted_ids = list(id_to_name.keys())
    sub = shares_df[shares_df["user_id"].isin(wanted_ids)].copy()
    piv = sub.pivot_table(index="expense_id", columns="user_id", values="owed_share", aggfunc="sum", fill_value=0.0)
    piv = piv.reindex(columns=wanted_ids, fill_value=0.0)  # fixed 2‑member order
    piv.columns = [f"{id_to_name.get(int(uid), str(uid))}_owed" for uid in piv.columns]
    out = out.merge(piv.reset_index(), on="expense_id", how="left")
    for display_name in id_to_name.values():
        c = f"{display_name}_owed"
        if c not in out.columns:
            out[c] = 0.0
        out[c] = pd.to_numeric(out[c], errors="coerce")
        out[c] = out[c].fillna(0.0).round(2)
    return out


def apply_custom_exclusions(raw_df: pd.DataFrame, shares_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Custom exclusions (from SPLITWISE_EXCLUDE_MONTHS and SPLITWISE_EXCLUDE_DESCRIPTIONS env vars):
      - Drop expenses whose description matches SPLITWISE_EXCLUDE_DESCRIPTIONS
        AND whose month is in SPLITWISE_EXCLUDE_MONTHS.
    
    Env vars:
      SPLITWISE_EXCLUDE_MONTHS: comma-separated months (e.g., "2025-10,2025-11")
      SPLITWISE_EXCLUDE_DESCRIPTIONS: comma-separated substrings to match case-insensitive (e.g., "פריז,trip")
    
    Applies to both Raw_Expenses and Raw_Shares/PerPerson_Month.
    """
    if raw_df is None or raw_df.empty:
        return raw_df, shares_df

    # Parse exclusion rules from environment
    exclude_months_str = os.getenv("SPLITWISE_EXCLUDE_MONTHS", "").strip()
    exclude_descs_str = os.getenv("SPLITWISE_EXCLUDE_DESCRIPTIONS", "").strip()

    if not exclude_months_str or not exclude_descs_str:
        # No exclusions configured
        return raw_df, shares_df

    months_to_drop = {m.strip() for m in exclude_months_str.split(",") if m.strip()}
    desc_patterns = [d.strip() for d in exclude_descs_str.split(",") if d.strip()]

    if not months_to_drop or not desc_patterns:
        return raw_df, shares_df

    month_series = raw_df.get("month").astype(str)
    desc_series = raw_df.get("description").astype(str)

    # rows to drop: month in months_to_drop AND description contains any pattern
    mask_drop = month_series.isin(months_to_drop)
    for pattern in desc_patterns:
        mask_drop = mask_drop & desc_series.str.contains(pattern, case=False, na=False, regex=False)

    # keep everything else
    mask_keep = ~mask_drop
    filtered_raw = raw_df[mask_keep].copy()

    # Keep only shares belonging to kept expense_ids
    if shares_df is not None and not shares_df.empty and "expense_id" in filtered_raw.columns:
        allowed_ids = set(filtered_raw["expense_id"].dropna().astype("Int64").tolist())
        filtered_shares = shares_df[shares_df["expense_id"].isin(allowed_ids)].copy()
    else:
        filtered_shares = shares_df

    return filtered_raw, filtered_shares


def build_pivots(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if df.empty:
        return df, df
    pivot = pd.pivot_table(
        df, index="month", columns="category", values="amount",
        aggfunc="sum", fill_value=0.0, sort=True
    ).sort_index()
    pivot["Total"] = pivot.sum(axis=1)
    mom = pivot.diff(); mom.columns = [f"Δ {c}" for c in mom.columns]
    pct = pivot.pct_change().replace([pd.NA, pd.NaT, float("inf"), float("-inf")], 0.0)
    pct.columns = [f"Δ% {c}" for c in pct.columns]
    mom_full = pd.concat([pivot, mom, pct], axis=1)
    return pivot, mom_full


def build_person_month_pivot(person_df: pd.DataFrame, restrict_to_ids: Optional[List[int]] = None) -> pd.DataFrame:
    if person_df.empty:
        return person_df
    df = person_df.copy()
    if restrict_to_ids:
        df = df[df["user_id"].isin(restrict_to_ids)].copy()
    # map to fixed Hebrew names
    df["person"] = df.apply(lambda r: MEMBER_ID_TO_NAME.get(int(r["user_id"]) if pd.notna(r["user_id"]) else -1, r["person"]), axis=1)
    pivot = pd.pivot_table(
        df, index="month", columns="person", values="owed_share",
        aggfunc="sum", fill_value=0.0, sort=True
    ).sort_index()
    # Only your two members as columns (in fixed order) + Total
    ordered_cols = [MEMBER_ID_TO_NAME[k] for k in MEMBER_ID_TO_NAME.keys() if MEMBER_ID_TO_NAME[k] in pivot.columns]
    pivot = pivot.reindex(columns=ordered_cols, fill_value=0.0)
    pivot["Total"] = pivot.sum(axis=1)
    return pivot


def compute_group_owes_from_simplified_debts(group: Group, id_to_name: Dict[int, str]) -> Dict[str, float]:
    """
    Get how much each tracked member OWES from the group's simplified debts.
    We only look at the 'from' side (debtor), and ignore what people are supposed to get back.
    """
    totals: Dict[str, float] = {name: 0.0 for name in id_to_name.values()}

    # Try SDK method getSimplifiedDebts()
    debts = None
    getter = getattr(group, "getSimplifiedDebts", None)
    if callable(getter):
        try:
            debts = getter()
        except Exception:
            debts = None

    # Fallback: raw attribute 'simplified_debts'
    if debts is None:
        debts = getattr(group, "simplified_debts", None)

    if not debts:
        return totals

    for d in debts:
        # Dict shape from API
        if isinstance(d, dict):
            from_id = d.get("from") or d.get("from_user") or d.get("from_user_id")
            amount = d.get("amount")
        else:
            # SDK object shape
            from_id = _safe_attr(d, "getFrom") or _safe_attr(d, "from_user") or _safe_attr(d, "getFromUser")
            if hasattr(d, "amount"):
                amount = getattr(d, "amount")
            else:
                amount = _safe_attr(d, "getAmount")

        try:
            from_id = int(from_id)
        except (TypeError, ValueError):
            continue

        val = _coerce_float(amount, 0.0)
        name = id_to_name.get(from_id)
        if name:
            totals[name] = totals.get(name, 0.0) + val

    return totals


def read_existing_raw(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        return pd.DataFrame()
    try:
        df = pd.read_excel(path, sheet_name="Raw_Expenses", engine="openpyxl")
        if "expense_id" in df.columns:
            df["expense_id"] = pd.to_numeric(df["expense_id"], errors="coerce").astype("Int64")
        for c in ("date","updated_at","month","description","category","currency"):
            if c in df.columns:
                df[c] = df[c].astype(str)
        if "amount" in df.columns:
            df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0.0)

        # --- Coalesce legacy *_paid / *_owed into canonical *_owes; drop *_x/_y variants ---
        for display_name in MEMBER_ID_TO_NAME.values():
            owes = f"{display_name}_owes"
            owed = f"{display_name}_owed"
            paid = f"{display_name}_paid"
            variants = [owed, paid, f"{owed}_x", f"{owed}_y", f"{paid}_x", f"{paid}_y", f"{owes}_x", f"{owes}_y"]

            # Create/refresh canonical _owes from any available source
            src = None
            if owes in df.columns:
                src = owes
            else:
                for cand in variants:
                    if cand in df.columns:
                        src = cand
                        break
            if src is not None and owes != src:
                df[owes] = pd.to_numeric(df[src], errors="coerce").fillna(0.0).round(2)

            # Drop legacy columns & suffix variants
            for cand in variants:
                if cand in df.columns and cand != owes:
                    df.drop(columns=[cand], inplace=True, errors="ignore")

            # Ensure final _owes exists
            if owes not in df.columns:
                df[owes] = 0.0
            df[owes] = pd.to_numeric(df[owes], errors="coerce").fillna(0.0).round(2)

        return df
    except Exception:
        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def upsert_on_expense_id(existing: pd.DataFrame, incoming: pd.DataFrame) -> pd.DataFrame:
    if existing.empty:
        return incoming.copy()
    ex = existing.set_index("expense_id", drop=False)
    inc = incoming.set_index("expense_id", drop=False)
    all_cols = sorted(set(ex.columns) | set(inc.columns))
    ex = ex.reindex(columns=all_cols)
    inc = inc.reindex(columns=all_cols)
    ex.update(inc)
    new_only = inc[~inc.index.isin(ex.index)]
    combined = pd.concat([ex, new_only], axis=0)
    combined = combined.reset_index(drop=True)
    sort_cols = [c for c in ["date","category","description"] if c in combined.columns]
    if sort_cols:
        combined = combined.sort_values(sort_cols, kind="mergesort")
    return combined


# --------------------------- Excel writing + charts ---------------------------
def write_excel(
    out_path: str,
    raw_df: pd.DataFrame,
    pivot: pd.DataFrame,
    mom: pd.DataFrame,
    person_pivot: pd.DataFrame,
    person_rows_df: pd.DataFrame | None = None,
) -> None:
    """
    Sheets:
      Raw_Expenses (includes <name>_owes columns),
      Raw_Shares (optional), Monthly_By_Category, PerPerson_Month,
      Charts:
        - Two 4x4 tiles (each member & total owed, centered, big font)
        - Total expenses per month
        - Per person owes by month (clustered, 2 bars per month)
        - Category expenses per month (dropdown for category)
        - Total by category (dropdown for month: 'All months' or specific month)
    """
    import time, tempfile, os
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, Border, Side

    tmp_fd, tmp_path = tempfile.mkstemp(prefix="splitwise_", suffix=".xlsx")
    os.close(tmp_fd)

    # Drop 'created_by' if ever present (focusing on owed splits)
    raw_to_write = raw_df.drop(columns=["created_by"], errors="ignore") if not raw_df.empty else raw_df

    with pd.ExcelWriter(tmp_path, engine="openpyxl") as xw:
        # --- Data sheets ---
        raw_to_write.to_excel(xw, sheet_name="Raw_Expenses", index=False)
        if person_rows_df is not None and not person_rows_df.empty:
            person_rows_df.to_excel(xw, sheet_name="Raw_Shares", index=False)
        if not pivot.empty:
            pivot.to_excel(xw, sheet_name="Monthly_By_Category")
        if not person_pivot.empty:
            person_pivot.to_excel(xw, sheet_name="PerPerson_Month")

        wb = xw.book

        # Fresh Charts sheet
        if "Charts" in wb.sheetnames:
            wb.remove(wb["Charts"])
        ws_charts = wb.create_sheet("Charts")

        # Make Charts the first sheet
        try:
            wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws_charts)))
        except Exception:
            pass

        # ---------- Summary tiles: each member & current group "owes" ----------
        from openpyxl.styles import Alignment, Font, Border, Side

        member_names = list(MEMBER_ID_TO_NAME.values())

        # Double border ("double coated margins")
        double_side = Side(style="double")
        tile_border = Border(
            left=double_side, right=double_side, top=double_side, bottom=double_side
        )

        # Layout: for each member we make 2x2 cells:
        #   row 1 (merged 2 columns): name
        #   row 2 (merged 2 columns): numeric value they owe now
        #
        # First member in B1:C2, second in F1:G2
        tile_positions = [("B", 1), ("F", 1)]

        for idx, name in enumerate(member_names[:2]):
            start_col_letter, start_row = tile_positions[idx]
            start_col_idx = ord(start_col_letter) - ord("A") + 1
            end_col_idx = start_col_idx + 1   # 2 columns wide

            name_row = start_row
            amount_row = start_row + 1

            # ----- Name row -----
            ws_charts.merge_cells(
                start_row=name_row,
                start_column=start_col_idx,
                end_row=name_row,
                end_column=end_col_idx,
            )
            name_cell = ws_charts.cell(row=name_row, column=start_col_idx)
            name_cell.value = name
            name_cell.style = "Calculation"
            name_cell.alignment = Alignment(horizontal="center", vertical="center")
            name_cell.font = Font(size=18, bold=True)

            # ----- Amount row (take value from group balance "owes") -----
            ws_charts.merge_cells(
                start_row=amount_row,
                start_column=start_col_idx,
                end_row=amount_row,
                end_column=end_col_idx,
            )
            amount_cell = ws_charts.cell(row=amount_row, column=start_col_idx)
            val = GLOBAL_MEMBER_OWES_SUM.get(name, 0.0)
            amount_cell.value = val
            amount_cell.style = "Calculation"
            amount_cell.number_format = u'₪ #,##0.00'
            amount_cell.alignment = Alignment(horizontal="center", vertical="center")
            amount_cell.font = Font(size=18, bold=True)

            # ----- Apply double border and keep center alignment on all 4 cells -----
            for r in (name_row, amount_row):
                for c in range(start_col_idx, end_col_idx + 1):
                    cell = ws_charts.cell(row=r, column=c)
                    cell.border = tile_border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


        # ---------- A) Total expenses per month (clustered bar) ----------
        if "Monthly_By_Category" in wb.sheetnames:
            ws_mbc = wb["Monthly_By_Category"]
            headers = [c.value for c in ws_mbc[1]]
            if headers and "Total" in headers:
                total_col = headers.index("Total") + 1
                max_row = ws_mbc.max_row

                data = Reference(ws_mbc, min_col=total_col, min_row=1, max_col=total_col, max_row=max_row)
                cats = Reference(ws_mbc, min_col=1, min_row=2, max_row=max_row)

                c = BarChart()
                c.type = "col"
                c.grouping = "clustered"
                c.title = "Total expenses per month"  # inside chart
                c.y_axis.title = "Amount"
                c.x_axis.title = "Month"
                c.y_axis.number_format = '#,##0.00'

                c.add_data(data, titles_from_data=True)
                c.set_categories(cats)

                c.dLbls = DataLabelList()
                c.dLbls.showVal = True
                c.dLbls.showSerName = False
                c.dLbls.showCatName = False
                c.dLbls.showLegendKey = False

                c.x_axis.delete = False
                c.y_axis.delete = False
                c.legend = None

                # Place below the 4x4 tiles
                ws_charts.add_chart(c, "A6")

        # ---------- B) Per person (top 2, clustered) ----------
        if "PerPerson_Month" in wb.sheetnames:
            ws_ppm = wb["PerPerson_Month"]
            headers = [c.value for c in ws_ppm[1]]
            max_row = ws_ppm.max_row

            # choose top 2 people by total (skip month/Total)
            person_cols = [(i + 1, h) for i, h in enumerate(headers) if h and h not in ("month", "Total")]
            totals_pp = []
            for col_idx, name in person_cols:
                s = 0.0
                for r in range(2, max_row + 1):
                    v = ws_ppm.cell(row=r, column=col_idx).value or 0
                    try:
                        s += float(v)
                    except Exception:
                        pass
                totals_pp.append((s, col_idx, name))
            totals_pp.sort(reverse=True)
            top2 = totals_pp[:2]

            if top2:
                c = BarChart()
                c.type = "col"
                c.grouping = "clustered"
                # title without "(top 2)"
                c.title = "Per person owes by month"
                c.y_axis.title = "Amount"
                c.x_axis.title = "Month"
                c.y_axis.number_format = '#,##0.00'

                cats = Reference(ws_ppm, min_col=1, min_row=2, max_row=max_row)

                for _, col_idx, _name in top2:
                    series = Reference(ws_ppm, min_col=col_idx, min_row=1, max_col=col_idx, max_row=max_row)
                    c.add_data(series, titles_from_data=True)

                c.set_categories(cats)

                c.dLbls = DataLabelList()
                c.dLbls.showVal = True
                c.dLbls.showSerName = False
                c.dLbls.showCatName = False
                c.dLbls.showLegendKey = False

                c.x_axis.delete = False
                c.y_axis.delete = False

                if len(top2) == 2:
                    c.legend.overlay = False
                    c.legend.position = "r"
                else:
                    c.legend = None

                # Move this chart down so there's comfortable space under the first chart
                # (was A18; now a bit lower to keep clear separation)
                ws_charts.add_chart(c, "A22")

        # ---------- C) Category-by-month (dropdown, single series) ----------
        if "Monthly_By_Category" in wb.sheetnames:
            ws_mbc = wb["Monthly_By_Category"]
            hdrs = [c.value for c in ws_mbc[1]]
            cats_only = [h for h in hdrs if h and h not in ("month", "Total")]

            if cats_only:
                # Helper sheet for selected category view
                if "CategoryView" in wb.sheetnames:
                    wb.remove(wb["CategoryView"])
                ws_cv = wb.create_sheet("CategoryView")
                ws_cv["A1"].value = "Month"
                ws_cv["B1"].value = "Selected Category"

                max_row = ws_mbc.max_row
                for r in range(2, max_row + 1):
                    # Month
                    ws_cv[f"A{r}"].value = f"='Monthly_By_Category'!A{r}"
                    # Value for selected category (with IFERROR so 'Choose Category' doesn't explode)
                    ws_cv[f"B{r}"].value = (
                        "=IFERROR("
                        "INDEX('Monthly_By_Category'!$1:$1048576,"
                        f"ROW(),MATCH(Charts!$L$20,'Monthly_By_Category'!$1:$1,0)),"
                        "NA())"
                    )

                # Dropdown cell adjacent to the category chart
                selector_cell = "L20"
                ws_charts[selector_cell].value = "Choose Category"

                # Hidden category list for dropdown (on Charts sheet, column Z)
                for i, name in enumerate(cats_only, start=1):
                    ws_charts[f"Z{i}"].value = name
                ws_charts.column_dimensions["Z"].hidden = True

                # Data validation: list from Z1:Z<n>
                dv = DataValidation(
                    type="list",
                    formula1=f"=Charts!$Z$1:$Z${len(cats_only)}",
                    allow_blank=True,
                )
                ws_charts.add_data_validation(dv)
                dv.add(ws_charts[selector_cell])

                # Chart pulling from CategoryView
                data = Reference(ws_cv, min_col=2, min_row=1, max_col=2, max_row=max_row)
                lab = Reference(ws_cv, min_col=1, min_row=2, max_row=max_row)

                c = BarChart()
                c.type = "col"
                c.grouping = "clustered"
                c.title = "Category expenses per month"
                c.y_axis.title = "Amount"
                c.x_axis.title = "Month"
                c.y_axis.number_format = '#,##0.00'

                c.add_data(data, titles_from_data=True)
                c.set_categories(lab)

                c.dLbls = DataLabelList()
                c.dLbls.showVal = True
                c.dLbls.showSerName = False
                c.dLbls.showCatName = False
                c.dLbls.showLegendKey = False

                c.x_axis.delete = False
                c.y_axis.delete = False
                c.legend = None

                # Place the chart near the selector (below it)
                ws_charts.add_chart(c, "L22")

        # ---------- D) Total by category (month dropdown: All months or specific month) ----------
        if "Monthly_By_Category" in wb.sheetnames:
            ws_mbc = wb["Monthly_By_Category"]
            headers = [c.value for c in ws_mbc[1]] or []
            cats_only = [h for h in headers if h and h not in ("month", "Total")]
            max_row = ws_mbc.max_row

            # Gather month labels from column A (row 2..)
            months = []
            for r in range(2, max_row + 1):
                val = ws_mbc.cell(row=r, column=1).value
                if val and val not in months:
                    months.append(val)

            if cats_only:
                # Helper sheet for totals
                if "CategoryTotals" in wb.sheetnames:
                    wb.remove(wb["CategoryTotals"])
                ws_ct = wb.create_sheet("CategoryTotals")
                ws_ct["A1"].value = "Category"
                ws_ct["B1"].value = "Total"

                # Put categories in col A
                for i, cat in enumerate(cats_only, start=2):
                    ws_ct[f"A{i}"].value = cat
                last_cat_row = len(cats_only) + 1  # last row index with a category

                # Month selector for this chart: Charts!L2
                month_selector_cell = "L2"
                ws_charts[month_selector_cell].value = "All months"

                # Hidden list of month options (All months + each real month) in Charts!Y
                all_month_options = ["All months"] + months
                for i, mval in enumerate(all_month_options, start=1):
                    ws_charts[f"Y{i}"].value = mval
                ws_charts.column_dimensions["Y"].hidden = True

                dv_month = DataValidation(
                    type="list",
                    formula1=f"=Charts!$Y$1:$Y${len(all_month_options)}",
                    allow_blank=True,
                )
                ws_charts.add_data_validation(dv_month)
                dv_month.add(ws_charts[month_selector_cell])

                # Fill CategoryTotals!B with formulas depending on selected month
                month_range = f"'Monthly_By_Category'!$A$2:$A${max_row}"
                for i, cat in enumerate(cats_only, start=2):
                    # find column index and letter in Monthly_By_Category
                    col_idx = headers.index(cat) + 1
                    col_letter = get_column_letter(col_idx)
                    data_range = f"'Monthly_By_Category'!${col_letter}$2:${col_letter}${max_row}"

                    sum_all = f"SUM({data_range})"
                    index_by_month = (
                        f"IFERROR("
                        f"INDEX({data_range}, MATCH(Charts!$L$2, {month_range}, 0)),"
                        "0)"
                    )

                    # If Charts!L2 = "All months" → SUM over all months
                    # else → value for the selected month
                    formula = f'=IF(Charts!$L$2="All months", {sum_all}, {index_by_month})'
                    ws_ct[f"B{i}"].value = formula

                # Build chart: total by category (depending on selected month)
                data = Reference(ws_ct, min_col=2, min_row=1, max_col=2, max_row=last_cat_row)
                cats_ref = Reference(ws_ct, min_col=1, min_row=2, max_row=last_cat_row)

                c = BarChart()
                c.type = "col"
                c.grouping = "clustered"
                c.title = "Total by category"
                c.y_axis.title = "Amount"
                c.x_axis.title = "Category"
                c.y_axis.number_format = '#,##0.00'

                c.add_data(data, titles_from_data=True)
                c.set_categories(cats_ref)

                c.dLbls = DataLabelList()
                c.dLbls.showVal = True
                c.dLbls.showSerName = False
                c.dLbls.showCatName = False
                c.dLbls.showLegendKey = False

                c.x_axis.delete = False
                c.y_axis.delete = False
                c.legend = None

                # Place chart; the month dropdown is at L2, chart just below/right
                ws_charts.add_chart(c, "L4")

    # Move temp file into place with retries (Windows-friendly)
    for _ in range(5):
        try:
            os.replace(tmp_path, out_path)
            return
        except PermissionError:
            time.sleep(1.2)
    base, ext = os.path.splitext(out_path)
    alt = f"{base}_new{ext}"
    try:
        os.replace(tmp_path, alt)
        print(f"[!] Target file was locked. Wrote to '{alt}' instead.")
    except PermissionError:
        print(f"[!] Still locked. Temporary file left at: {tmp_path}")



# ---------- CLI ----------
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Splitwise → Excel (all‑history, incremental upsert + per‑person paid)")
    p.add_argument("--out", help="Output Excel path (.xlsx). Default: splitwise_group_<id>_all_history.xlsx")
    p.add_argument("--start", help="Override start month (YYYY-MM). Default: SPLITWISE_FIRST_MONTH env var or 2008-01")
    p.add_argument("--end", help="Override end month (YYYY-MM). Default: current month")
    p.add_argument("--group-id", type=int, help="Override group id. Default: SPLITWISE_GROUP_ID env var")
    p.add_argument("--no-raw-shares", action="store_true", help="Don't write Raw_Shares sheet")
    p.add_argument("--debug-shares", action="store_true", help="Print a short sanity summary of parsed per-user shares")
    return p.parse_args()


def resolve_token(token_arg: Optional[str]) -> Dict:
    if not token_arg:
        raise SystemExit("[!] Provide OAuth token JSON via SPLITWISE_ACCESS_TOKEN_JSON (GitHub Secret)")
    token_arg = token_arg.strip()
    if token_arg.startswith("@"):
        path = token_arg[1:]
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    try:
        return json.loads(token_arg)
    except json.JSONDecodeError:
        raise SystemExit("[!] Could not parse token JSON. If in a file, pass @/path/to/token.json")


def main():
    load_dotenv()
    args = parse_args()

    if not DEFAULT_GROUP_ID:
        raise SystemExit("[!] SPLITWISE_GROUP_ID not set in environment. Provide via .env or --group-id")

    group_id = args.group_id or DEFAULT_GROUP_ID
    start_ym = args.start or FIRST_MONTH
    end_ym = args.end or ym_today()
    out_path = args.out or OUTPUT_FILE_TEMPLATE.format(group_id=group_id)

    client_id = os.getenv("SPLITWISE_CLIENT_ID")
    client_secret = os.getenv("SPLITWISE_CLIENT_SECRET")
    token_json_raw = os.getenv("SPLITWISE_ACCESS_TOKEN_JSON")
    if not client_id or not client_secret:
        raise SystemExit("[!] Missing SPLITWISE_CLIENT_ID or SPLITWISE_CLIENT_SECRET (set as GitHub Secrets).")
    token = resolve_token(token_json_raw)

    sw = make_client(client_id, client_secret, token)
    group = fetch_group(sw, group_id)
    print(f"Group '{group.getName()}' (id={group_id})")

    # Fill per-member "owes" summary from group balances (simplified debts)
    GLOBAL_MEMBER_OWES_SUM.clear()
    GLOBAL_MEMBER_OWES_SUM.update(
        compute_group_owes_from_simplified_debts(group, MEMBER_ID_TO_NAME)
    )

    print(f"Fetching expenses from {start_ym} to {end_ym}…")
    expenses = fetch_expenses_all_history(sw, group_id, start_ym, end_ym)
    print(f"Fetched {len(expenses)} expenses from API")

    # Normalize flat expenses and long per-user owed rows (from users[].owed_share)
    incoming_df = normalize_expenses(expenses)
    person_rows_df = normalize_person_shares(expenses)

    # Exclude 'פריז' expenses in 2025-11
    before_raw = len(incoming_df)
    before_shares = 0 if person_rows_df is None else len(person_rows_df)
    incoming_df, person_rows_df = apply_custom_exclusions(incoming_df, person_rows_df)
    removed_raw = before_raw - len(incoming_df)
    removed_shares = before_shares - (0 if person_rows_df is None else len(person_rows_df))
    if removed_raw or removed_shares:
        print(f"Applied exclusions → removed {removed_raw} expenses and {removed_shares} share rows.")

    if args.debug_shares:
        print(
            f"[debug] shares rows: {len(person_rows_df)} "
            f"unique expenses in shares: {person_rows_df['expense_id'].nunique() if not person_rows_df.empty else 0}"
        )
        if not person_rows_df.empty:
            print("[debug] sample shares:")
            print(person_rows_df.head(6).to_string(index=False))

    # Upsert raw (by expense only) to keep idempotency
    existing_df = read_existing_raw(out_path)
    if not existing_df.empty:
        print(f"Found existing workbook with {len(existing_df)} rows in Raw_Expenses")
    merged_df = upsert_on_expense_id(existing_df, incoming_df)

    # Expand raw with two owed columns for your two members (rounded to 2 decimals)
    raw_wide = widen_raw_with_member_owed(merged_df, person_rows_df, MEMBER_ID_TO_NAME)

    # Category and per-person monthly pivots (per-person limited to your two members)
    pivot, mom = build_pivots(raw_wide)
    person_pivot = build_person_month_pivot(
        person_rows_df,
        restrict_to_ids=list(MEMBER_ID_TO_NAME.keys())
    )

    # Write workbook (Raw_Shares optional)
    shares_out = None if args.no_raw_shares else person_rows_df
    write_excel(out_path, raw_wide, pivot, mom, person_pivot, shares_out)

    print(f"✅ Export complete → {out_path}")
    if not raw_wide.empty and raw_wide.get("currency", pd.Series()).nunique() > 1:
        print("[!] Multiple currencies detected. Amounts are not converted in this export.")



if __name__ == "__main__":
    main()
